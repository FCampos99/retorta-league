#!/usr/bin/env python3
"""Retorta League — servidor local.
Actualiza o Excel quando reportas um jogo e regenera o website.
Uso: python3 server.py
"""

import base64, json, os, re, socket, subprocess
from http.server import HTTPServer, BaseHTTPRequestHandler
from datetime import datetime
import openpyxl

REPO_DIR   = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(REPO_DIR, 'Retorta-League.xlsx')
HTML_PATH  = os.path.join(REPO_DIR, 'index.html')
PORT       = 8765

SEASON_SHEET_MAP = {
    'UCL_SS':     'UCL SS',
    'UCL_WS':     'UCL WS',
    'UCL_SS_2ed': 'UCL SS 2ed',
    'UCL_WS_2ed': 'UCL WS 2ed',
    'UCL_SS_3ed': 'UCL SS 3ed',
}

SEASONS_META = [
    #  sheet_name     tipo             edicao          ano       inicio        fim          atual  factor_rows
    ('UCL SS',     'Época de Verão',   '1.ª Edição', '2024',    '18/03/2024', '29/07/2024', False, (36, 54)),
    ('UCL WS',     'Época de Inverno', '1.ª Edição', '2024-25', '05/09/2024', '28/01/2025', False, (40, 57)),
    ('UCL SS 2ed', 'Época de Verão',   '2.ª Edição', '2025',    '06/03/2025', '07/07/2025', False, (55, 73)),
    ('UCL WS 2ed', 'Época de Inverno', '2.ª Edição', '2025-26', '08/09/2025', '19/01/2026', False, (42, 60)),
    ('UCL SS 3ed', 'Época de Verão',   '3.ª Edição', '2026',    '10/03/2026', '14/07/2026', True,  (29, 47)),
]


# ── HTTP Handler ────────────────────────────────────────────────────────────

class Handler(BaseHTTPRequestHandler):

    def do_OPTIONS(self):
        self._ok_cors()
        self.end_headers()

    def do_GET(self):
        if self.path == '/ping':
            self._ok_cors()
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(b'{"ok":true}')

    def do_POST(self):
        length = int(self.headers.get('Content-Length', 0))
        body   = self.rfile.read(length)
        try:
            data = json.loads(body)
            if self.path == '/report':
                result = process_report(data)
            elif self.path == '/upload-capa':
                result = save_capa(data)
            else:
                result = {'error': 'unknown endpoint'}
            self._ok_cors()
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps(result, ensure_ascii=False).encode())
        except Exception as e:
            self.send_response(500)
            self._cors_headers()
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'error': str(e)}, ensure_ascii=False).encode())

    def _ok_cors(self):
        self.send_response(200)
        self._cors_headers()

    def _cors_headers(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, GET, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')

    def log_message(self, fmt, *args):
        ts = datetime.now().strftime('%H:%M:%S')
        print(f'[{ts}] {fmt % args}')


# ── Excel update ────────────────────────────────────────────────────────────

def parse_pt_date(s):
    d, m, y = s.split('/')
    return datetime(int(y), int(m), int(d))

def process_report(data):
    """
    data = {
      season_id : 'UCL_SS_3ed',
      date      : '02/06/2026',   # DD/MM/YYYY
      winners   : ['Francisco Campos', ...],
      losers    : ['João Capela', ...],
    }
    """
    season_id = data['season_id']
    date_str  = data['date']
    winners   = set(data['winners'])
    losers    = set(data['losers'])

    sheet_name  = SEASON_SHEET_MAP[season_id]
    target_date = parse_pt_date(date_str)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[sheet_name]

    # Find the column whose header matches the target date
    date_col = None
    for col in range(12, ws.max_column + 2):
        val = ws.cell(row=2, column=col).value
        if isinstance(val, datetime) and val.date() == target_date.date():
            date_col = col
            break

    if date_col is None:
        raise ValueError(f'Jornada {date_str} não encontrada na época {season_id}')

    # Walk player rows (col 1 = rank, col 2 = name)
    updated   = []
    not_found = list(winners | losers)

    for row in range(3, ws.max_row + 1):
        rank = ws.cell(row=row, column=1).value
        name = ws.cell(row=row, column=2).value
        if not isinstance(rank, (int, float)) or not isinstance(name, str):
            continue
        if name in winners:
            ws.cell(row=row, column=date_col).value = 'W'
            updated.append(f'{name}: W')
            if name in not_found: not_found.remove(name)
        elif name in losers:
            ws.cell(row=row, column=date_col).value = 'L'
            updated.append(f'{name}: L')
            if name in not_found: not_found.remove(name)

    wb.save(EXCEL_PATH)
    print(f'  Excel guardado — {len(updated)} células actualizadas')

    regenerate_html()
    git_publish(date_str)

    return {'ok': True, 'updated': updated, 'not_found': not_found}


# ── Website regeneration ────────────────────────────────────────────────────

def extract_seasons():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    seasons = []

    for sheet_name, tipo, edicao, ano, inicio, fim, atual, factor_rows in SEASONS_META:
        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Read factor lookup table directly from the Excel (col B = jogos, col C = factor)
        factor_table = {}
        for r in range(factor_rows[0], factor_rows[1] + 1):
            key = ws.cell(row=r, column=2).value
            val = ws.cell(row=r, column=3).value
            if isinstance(key, (int, float)) and isinstance(val, (int, float)):
                factor_table[int(key)] = val

        match_dates = []
        for col in rows[1][11:]:
            if isinstance(col, datetime):
                match_dates.append(col.strftime('%d/%m/%Y'))
            elif col is not None:
                match_dates.append(str(col))

        players = []
        for row in rows[2:]:
            if row[0] is None or row[1] is None:
                continue
            if not isinstance(row[0], (int, float)):
                continue

            results = [row[11 + i] if 11 + i < len(row) and row[11 + i] in ('W', 'L') else None
                       for i in range(len(match_dates))]

            # Computar wins/losses/games dos dados W/L brutos (fiável mesmo após updates)
            vitorias = results.count('W')
            derrotas = results.count('L')
            jogos    = vitorias + derrotas
            if jogos == 0:
                continue

            win_rate = round(vitorias / jogos, 4)

            # mediaPonderada: usar o valor cacheado do Excel (calculado pelo próprio Excel)
            # Se não estiver disponível (célula de fórmula não recalculada), calcular com lookup table
            cached_media = row[9] if len(row) > 9 and isinstance(row[9], (int, float)) else None
            if cached_media is not None:
                media = round(cached_media, 4)
            else:
                factor = factor_table.get(jogos, 0)
                media  = round(win_rate * factor, 4)

            # rank: usar o do Excel directamente (não re-calcular)
            players.append({
                'rank': int(row[0]),
                'nome': row[1],
                'alcunha': row[2] or row[1],
                'jogos': jogos,
                'vitorias': vitorias,
                'derrotas': derrotas,
                'winRate': win_rate,
                'mediaPonderada': media,
                'resultados': results,
            })

        # Preservar a ordem das linhas do Excel (que reflecte o ranking definido no Excel)
        # só reordenar por mediaPonderada quando o rank do Excel não está disponível
        players.sort(key=lambda p: p['rank'])

        seasons.append({
            'id': sheet_name.replace(' ', '_'),
            'label': f'{tipo} · {edicao}',
            'tipo': tipo,
            'edicao': edicao,
            'ano': ano,
            'inicio': inicio,
            'fim': fim,
            'atual': atual,
            'datas': match_dates,
            'jogadores': players,
        })

    return seasons


def load_capas_index():
    idx_path = os.path.join(REPO_DIR, 'capas', 'index.json')
    try:
        return json.load(open(idx_path, encoding='utf-8'))
    except Exception:
        return []


def save_capa(data):
    """Save uploaded photo to capas/ folder and push to GitHub."""
    capas_dir = os.path.join(REPO_DIR, 'capas')
    os.makedirs(capas_dir, exist_ok=True)

    # Decode base64 image
    match = re.match(r'data:image/(\w+);base64,(.*)', data['src'], re.DOTALL)
    if not match:
        raise ValueError('Formato de imagem inválido')
    ext      = 'jpg' if match.group(1) in ('jpeg', 'jpg') else match.group(1)
    img_data = base64.b64decode(match.group(2))

    filename = datetime.now().strftime(f'%Y-%m-%d_%H%M.{ext}')
    img_path = os.path.join(capas_dir, filename)
    open(img_path, 'wb').write(img_data)

    # Update index.json
    idx_path = os.path.join(capas_dir, 'index.json')
    index    = load_capas_index()
    index.append({'filename': filename, 'date': data.get('date', datetime.now().isoformat()), 'caption': data.get('caption', '')})
    json.dump(index, open(idx_path, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)

    # Regenerate HTML so the new capa appears even offline
    regenerate_html()

    # Push to GitHub
    try:
        subprocess.run(['git', 'add', f'capas/{filename}', 'capas/index.json', 'index.html'], cwd=REPO_DIR, check=True, capture_output=True)
        subprocess.run(['git', 'commit', '-m', f'Capa: {filename}'], cwd=REPO_DIR, check=True, capture_output=True)
        subprocess.run(['git', 'push'], cwd=REPO_DIR, check=True, capture_output=True)
        print(f'  ✅ Capa publicada: {filename}')
    except subprocess.CalledProcessError as e:
        print(f'  ⚠️  Git push falhou: {e.stderr.decode() if e.stderr else e}')

    return {'ok': True, 'filename': filename}


def regenerate_html():
    seasons   = extract_seasons()
    capas     = load_capas_index()
    seasons_json = json.dumps(seasons, ensure_ascii=False)
    capas_json   = json.dumps(capas,   ensure_ascii=False)

    html = open(HTML_PATH, encoding='utf-8').read()
    html = re.sub(
        r'/\* RETORTA_DATA_BEGIN \*/.*?/\* RETORTA_DATA_END \*/',
        f'/* RETORTA_DATA_BEGIN */{seasons_json}/* RETORTA_DATA_END */',
        html, flags=re.DOTALL
    )
    html = re.sub(
        r'/\* CAPAS_DATA_BEGIN \*/.*?/\* CAPAS_DATA_END \*/',
        f'/* CAPAS_DATA_BEGIN */{capas_json}/* CAPAS_DATA_END */',
        html, flags=re.DOTALL
    )
    open(HTML_PATH, 'w', encoding='utf-8').write(html)
    print(f'  index.html regenerado — {len(seasons)} épocas, {len(capas)} capas')


def git_publish(date_str):
    try:
        subprocess.run(['git', 'add', 'index.html', 'Retorta-League.xlsx'], cwd=REPO_DIR, check=True, capture_output=True)
        subprocess.run(['git', 'commit', '-m', f'Jornada {date_str}'], cwd=REPO_DIR, check=True, capture_output=True)
        subprocess.run(['git', 'push'], cwd=REPO_DIR, check=True, capture_output=True)
        print(f'  ✅ Publicado em https://fcampos99.github.io/retorta-league/')
    except subprocess.CalledProcessError as e:
        print(f'  ⚠️  Git push falhou: {e.stderr.decode() if e.stderr else e}')


# ── Entry point ─────────────────────────────────────────────────────────────

if __name__ == '__main__':
    print('═' * 50)
    print('  Retorta League — Servidor Local')
    print(f'  Excel  : {EXCEL_PATH}')
    print(f'  Website: {HTML_PATH}')
    print('═' * 50)
    print('  A gerar dados iniciais...')
    regenerate_html()

    server = HTTPServer(('localhost', PORT), Handler)
    server.socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
    print(f'  Servidor a correr em http://localhost:{PORT}')
    print('  Abre o index.html e reporta jogos!')
    print('  Ctrl+C para parar.')
    print('═' * 50)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print('\n  Servidor parado.')
